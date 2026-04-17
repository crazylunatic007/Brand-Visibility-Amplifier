import io
import csv
from models.schemas import GeneratedQuery, QueryIntent, QueryType


def parse_uploaded_queries(uploaded_file) -> list[GeneratedQuery]:
    """Parse queries from CSV, TXT, or XLSX files.

    Supported formats:
    - TXT: One query per line
    - CSV: First column is the query text. Optional columns: intent, query_type, rationale
    - XLSX: Same as CSV but in Excel format
    """
    filename = uploaded_file.name.lower()

    if filename.endswith(".txt"):
        return _parse_txt(uploaded_file)
    elif filename.endswith(".csv"):
        return _parse_csv(uploaded_file)
    elif filename.endswith(".xlsx"):
        return _parse_xlsx(uploaded_file)
    else:
        raise ValueError(f"Unsupported file type: {filename}. Use CSV, TXT, or XLSX.")


def _parse_txt(uploaded_file) -> list[GeneratedQuery]:
    """One query per line."""
    content = uploaded_file.read().decode("utf-8", errors="ignore")
    lines = [line.strip() for line in content.splitlines() if line.strip()]
    return [_make_query(text) for text in lines]


def _parse_csv(uploaded_file) -> list[GeneratedQuery]:
    """CSV with columns: query_text (required), intent, query_type, rationale (optional)."""
    content = uploaded_file.read().decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(content))

    # Also handle CSVs without headers (just a list of queries)
    if reader.fieldnames is None:
        lines = [line.strip() for line in content.splitlines() if line.strip()]
        return [_make_query(text) for text in lines]

    # Check if first column header looks like a query (no standard header)
    has_query_col = any(
        col.lower() in ("query", "query_text", "keyword", "search_query", "queries")
        for col in (reader.fieldnames or [])
    )

    results = []
    if has_query_col:
        for row in reader:
            query_text = (
                row.get("query")
                or row.get("query_text")
                or row.get("keyword")
                or row.get("search_query")
                or row.get("queries")
                or ""
            ).strip()
            if not query_text:
                continue
            intent = _safe_intent(row.get("intent", ""))
            query_type = _safe_query_type(row.get("query_type", "") or row.get("type", ""))
            rationale = row.get("rationale", "Uploaded by user")
            results.append(
                GeneratedQuery(
                    query_text=query_text,
                    intent=intent,
                    query_type=query_type,
                    rationale=rationale or "Uploaded by user",
                )
            )
    else:
        # No recognized header — treat first column as query text
        content = uploaded_file.getvalue().decode("utf-8", errors="ignore") if hasattr(uploaded_file, 'getvalue') else content
        for line in content.splitlines():
            parts = line.split(",")
            text = parts[0].strip().strip('"').strip("'")
            if text and text.lower() not in ("query", "query_text", "keyword"):
                results.append(_make_query(text))

    return results


def _parse_xlsx(uploaded_file) -> list[GeneratedQuery]:
    """XLSX with columns: query_text (required), intent, query_type, rationale (optional)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(uploaded_file.read()), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Check if first row is a header
    header = [str(cell).lower().strip() if cell else "" for cell in rows[0]]
    query_col = None
    intent_col = None
    type_col = None
    rationale_col = None

    for i, h in enumerate(header):
        if h in ("query", "query_text", "keyword", "search_query", "queries"):
            query_col = i
        elif h == "intent":
            intent_col = i
        elif h in ("query_type", "type"):
            type_col = i
        elif h == "rationale":
            rationale_col = i

    results = []
    start_row = 1 if query_col is not None else 0
    if query_col is None:
        query_col = 0  # Default to first column

    for row in rows[start_row:]:
        text = str(row[query_col]).strip() if row[query_col] else ""
        if not text:
            continue
        intent = _safe_intent(str(row[intent_col]).strip() if intent_col and row[intent_col] else "")
        query_type = _safe_query_type(str(row[type_col]).strip() if type_col and row[type_col] else "")
        rationale = str(row[rationale_col]).strip() if rationale_col and row[rationale_col] else "Uploaded by user"
        results.append(
            GeneratedQuery(
                query_text=text,
                intent=intent,
                query_type=query_type,
                rationale=rationale,
            )
        )

    wb.close()
    return results


def _make_query(text: str) -> GeneratedQuery:
    """Create a GeneratedQuery with defaults for an uploaded query."""
    return GeneratedQuery(
        query_text=text,
        intent=QueryIntent.INFORMATIONAL,
        query_type=QueryType.SPECIFICATION,
        rationale="Uploaded by user",
    )


def _safe_intent(value: str) -> QueryIntent:
    """Try to match a string to QueryIntent, default to INFORMATIONAL."""
    value = value.lower().strip()
    for member in QueryIntent:
        if member.value == value:
            return member
    return QueryIntent.INFORMATIONAL


def _safe_query_type(value: str) -> QueryType:
    """Try to match a string to QueryType, default to SPECIFICATION."""
    value = value.lower().strip()
    for member in QueryType:
        if member.value == value:
            return member
    return QueryType.SPECIFICATION
